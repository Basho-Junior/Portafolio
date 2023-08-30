import datetime
import os
import time

import requests
from bs4 import BeautifulSoup
import tkinter as tk
import tkinter.messagebox as messagebox
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import urllib3
import threading

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configuración del correo electrónico
sender_email = "impresoraxerox979@gmail.com"
sender_password = "alezbbjbhxvlzoji"
receiver_email = "juniorhernandez38@hotmail.com"
smtp_server = "smtp.gmail.com"
smtp_port = 587
alert_windows = []

# Agrega esta variable global al inicio del programa
ips_window = None

# Leer las IPs desde el archivo ips.txt y asignarlas a la lista urls
with open("Impresoras.txt", "r") as file:
    impresora_lines = file.readlines()
    impresoras = [line.strip() for line in impresora_lines]

urls = [(f"https://{impresora}/stat/consumables.php", impresora) for impresora in impresoras]
urls2 = [f"https://{impresora}/stat/welcome.php?tab=status" for impresora in impresoras]

# Lista para almacenar las alertas previamente guardadas en el archivo TXT
existing_alerts = []

def create_new_window(ip, data):
    new_window = tk.Toplevel()
    new_window.title("Alerta!")
    new_window.geometry("300x150")

    ip_label = tk.Label(new_window, text=f"IP: {ip}")
    ip_label.pack()

    for component, status, life_remaining in data:
        component_label = tk.Label(new_window, text=f"Component: {component}")
        component_label.pack()

        status_label = tk.Label(new_window, text=f"Status: {status}")
        status_label.pack()

        life_remaining_label = tk.Label(new_window, text=f"Life Remaining: {life_remaining}")
        life_remaining_label.pack()

    alert_windows.append(new_window)


def send_email(subject, message):
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = receiver_email
    if "172.16.25" in subject:
        subject = "LAS CARRERAS"
    elif "192.168.17" in subject:
        subject = "LA OTRA BANDA"
    msg["Subject"] = subject
    msg.attach(MIMEText(message, "plain"))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
    except Exception as e:
        print(f"Error sending email: {e}")

    save_alert_to_txt(subject, message)

# Agregar la función para guardar la alerta en el archivo TXT
def save_alert_to_txt(alert_key, alert_message):
    filename = "alertas.txt"
    with open(filename, "a") as file:
        file.write(f"{alert_key}\n{alert_message}\n\n")

# Función para leer las alertas previamente guardadas desde el archivo TXT
def read_existing_alerts():
    filename = "alertas.txt"
    if os.path.exists(filename):
        with open(filename, "r") as file:
            lines = file.read().split("\n\n")
            for i in range(0, len(lines), 2):
                alert_key = lines[i].strip()
                existing_alerts.append(alert_key)

read_existing_alerts()


def show_alert(ip, data):
    # Construir el mensaje de alerta
    alert_message = f"ALERT: IP: {ip}\nComponents:\n"
    for component, status, life_remaining in data:
        alert_message += f"Component: {component}\n"
        alert_message += f"Status: {status}\n"
        alert_message += f"Life Remaining: {life_remaining}\n\n"

    # Mostrar la alerta en la consola
    print(alert_message)

    # Verificar si la alerta ya existe en el archivo TXT
    alert_key = f"{ip}-{data[0][0]}-{data[0][1]}"  # Clave única para la alerta
    if alert_key in existing_alerts:
        print("Alerta ya existe en el archivo TXT, no se enviará por correo electrónico.")
        return

    # Enviar la alerta por correo electrónico
    subject = f"ALERT: Printer at IP {ip}"
    send_email(subject, alert_message)

    # Mostrar la alerta en la ventana emergente
    create_new_window(ip, data)

    # Guardar la alerta en la lista de alertas previamente guardadas
    existing_alerts.append(alert_key)

    # Cerrar todas las ventanas de alerta abiertas
    close_alert_windows()


def show_alert_message(ip, component, status, life_remaining):
    if status == "missing":
        alert_message = f"Maquina detenida por el componente {component} con el estado {status}"
        # Mostrar la alerta en la consola
        print(alert_message)

        # Verificar si la alerta ya existe en el archivo TXT
        alert_key = f"{ip}-{component}-{status}"  # Clave única para la alerta
        if alert_key in existing_alerts:
            print("Alerta ya existe en el archivo TXT, no se mostrará en la ventana emergente.")
            return

        # Mostrar la alerta en la ventana emergente
        messagebox.showinfo("Alerta", alert_message)

        # Guardar la alerta en la lista de alertas previamente guardadas
        existing_alerts.append(alert_key)

        close_alert_windows()


def scrape_data(url):
    response = requests.get(url, verify=False)
    soup = BeautifulSoup(response.text, 'html.parser')
    table = soup.find('table', {'id': 'tableBilling'})
    rows = table.find_all('tr')

    data = []
    for row in rows:
        cells = row.find_all('td')
        if len(cells) == 2:
            data.append((cells[0].text.strip(), cells[1].text.strip()))

    return data


def format_table(sheet):
    # Formato de la tabla
    for column_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        column = sheet[column_letter]
        column[0].font = Font(bold=True)
        column[0].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        column[0].alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar ancho de columnas
    for column_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        sheet.column_dimensions[column_letter].width = 15

    # Ajustar alto de filas
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Ajustar tamaño de fuente
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width


def save_data(ip, data, component_data):
    filename = "Reporte.xlsx"
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
    else:
        workbook = openpyxl.load_workbook(filename)

    if ip in workbook.sheetnames:
        # Si la hoja ya existe, eliminarla antes de agregarla nuevamente
        workbook.remove(workbook[ip])

    sheet = workbook.create_sheet(ip)

    sheet['A1'] = 'IP'
    sheet['B1'] = 'Número de serie del dispositivo'
    sheet['C1'] = 'Impresiones en negro'
    sheet['D1'] = 'Impresiones en color'
    sheet['E1'] = 'Total de impresiones'
    sheet['F1'] = 'Impresiones gran tamaño en negro'
    sheet['G1'] = 'Impresiones gran tamaño en color'
    sheet['H1'] = 'Componente'
    sheet['I1'] = 'Estado'
    sheet['J1'] = 'Porcentaje'

    sheet['A2'] = ip
    sheet['B2'] = data[0][0].replace('Device Serial Number: ', '')
    sheet['C2'] = data[1][1]
    sheet['D2'] = data[2][1]
    sheet['E2'] = data[3][1]
    sheet['F2'] = data[4][1]
    sheet['G2'] = data[5][1]

    row = 2
    for component, status, life_remaining in component_data:
        sheet[f'H{row}'] = component
        sheet[f'I{row}'] = status
        sheet[f'J{row}'] = life_remaining
        row += 1

    format_table(sheet)
    workbook.save(filename)


def show_main_window():
    window.update()
    window.deiconify()


ips_window = None


def show_printer_ips(printer_ips, main_window):
    global ips_window

    if ips_window is not None and ips_window.winfo_exists():
        return

    ips_window = tk.Toplevel(main_window)
    ips_window.title("Printer IPs")
    ips_window.geometry("200x300")

    ips_label = tk.Label(ips_window, text="Printer IPs:")
    ips_label.pack()

    checkbox_vars = []
    for ip in printer_ips:
        ip_address = ip.split('/')[2]
        var = tk.BooleanVar()
        checkbox = tk.Checkbutton(ips_window, text=ip_address, variable=var)
        checkbox.pack(pady=5)
        checkbox_vars.append((ip, var))

    generate_button = tk.Button(ips_window, text="Generar Excel",
                                command=lambda: handle_generate_button(checkbox_vars, ips_window))
    generate_button.pack(pady=10)

    main_window.withdraw()
    ips_window.protocol("WM_DELETE_WINDOW", close_program)


def handle_generate_button(checkbox_vars, main_window):
    selected_ips = [ip for ip, var in checkbox_vars if var.get()]
    for ip in selected_ips:
        handle_ip_button(ip, main_window)

    close_ips_window()
    printer_ips2 = urls2.copy()
    show_printer_ips(printer_ips2, window)


def close_ips_window():
    global ips_window  # Accede a la variable global

    if ips_window is not None:
        ips_window.destroy()
        ips_window = None  # Reinicia la variable global


def extract_ip_from_url(url):
    # Utilizamos una expresión regular para extraer la dirección IP de la URL
    match = re.search(r'https://([\d.]+)/', url)
    if match:
        return match.group(1)
    else:
        raise ValueError('No se pudo extraer la dirección IP de la URL proporcionada.')


def handle_ip_button(ip, main_window):
    data = scrape_data(ip)

    ip_address = extract_ip_from_url(ip)
    urls = [
        (f'https://{ip_address}/stat/consumables.php', ip_address)
    ]

    printer_ips = []
    component_data = []
    for url, ips in urls:
        try:
            component_data.extend(get_toner_data(url, ips, printer_ips))
        except Exception as e:
            print(f"Error handling printer {ip}: {e}")

    save_data(ip.split('/')[2], data, component_data)
    show_alert_message(ip, data[0][0], data[0][1], data[1][1])


def get_toner_data(url, ip, printer_ips):
    component_data = []
    try:
        response = requests.get(url, verify=False)
        soup = BeautifulSoup(response.text, 'html.parser')

        print(f'URL: {url}\tIP: {ip}')

        toner_table = soup.find('div', class_='boundingBox normalSpacing noBottom')
        toner_rows = toner_table.find_all('tr')

        print('Toner Cartridges')
        print('  Component\t\t Status\t Life Remaining')
        data = []
        for row in toner_rows[1:]:
            columns = row.find_all('td')
            if len(columns) >= 4:  # Verificar que haya suficientes columnas
                component = columns[1].text.strip()
                status = columns[2].text.strip()
                life_remaining_px = columns[3].find('div', class_='horizBarFilledDiv')['style'].split('width: ')[
                    1].split(';')[0]
                life_remaining = int(life_remaining_px.split('px')[0]) / 2
                print(f'  {component}\t {status}\t {life_remaining}')
                component_data.append((component, status, life_remaining))
                if life_remaining < 20 and component != 'Waste Toner Container':
                    data.append((component, status, life_remaining))
                    if status == "missing":
                        show_alert_message(ip, component, status, life_remaining)
            else:
                print('')

        drum_table = soup.find_all('div', class_='boundingBox normalSpacing noBottom')[1]
        drum_rows = drum_table.find_all('tr')

        print('Drum Cartridges')
        print('  Component\t\t Status\t Life Remaining')
        for row in drum_rows[1:]:
            columns = row.find_all('td')
            if len(columns) >= 4:  # Verificar que haya suficientes columnas
                component = columns[1].text.strip()
                status = columns[2].text.strip()
                life_remaining_px = columns[3].find('div', class_='horizBarFilledDiv')['style'].split('width: ')[
                    1].split(';')[0]
                life_remaining = int(life_remaining_px.split('px')[0]) / 2
                print(f'  {component}\t {status}\t {life_remaining}')
                component_data.append((component, status, life_remaining))
                if life_remaining < 20 and component != 'Waste Toner Container':
                    data.append((component, status, life_remaining))
                    if status == "missing":
                        show_alert_message(ip, component, status, life_remaining)
            else:
                print('')

        waste_toner_parts = soup.find_all('div', class_='boundingBox normalSpacing noBottom')[2]
        waste_toner_rows = waste_toner_parts.find_all('tr')
        print('Waste Toner Container')
        print('  Component\t\t Status')
        for row in waste_toner_rows[1:]:
            columns = row.find_all('td')
            if len(columns) >= 4:  # Verificar que haya suficientes columnas
                component = columns[1].text.strip()
                status = columns[2].text.strip()
                print(f'  {component}\t {status}')
                component_data.append((component, status, ""))
                if status == 'Reorder':
                    data.append((component, status, ""))
            else:
                print('')

        serviceable_parts = soup.find_all('div', class_='boundingBox normalSpacing noBottom')[3]
        serviceable_rows = serviceable_parts.find_all('tr')

        print('Serviceable Parts')
        print('  Component\t\t Status\t Life Remaining')
        for row in serviceable_rows[1:]:
            columns = row.find_all('td')
            if len(columns) >= 4:  # Verificar que haya suficientes columnas
                component = columns[1].text.strip()
                status = columns[2].text.strip()
                life_remaining_px = columns[3].find('div', class_='horizBarFilledDiv')['style'].split('width: ')[
                    1].split(';')[0]
                life_remaining = int(life_remaining_px.split('px')[0]) / 2
                print(f'  {component}\t {status}\t {life_remaining}')
                component_data.append((component, status, life_remaining))
                if life_remaining < 20 and component != 'Waste Toner Container':
                    data.append((component, status, life_remaining))
                    if status == "missing":
                        show_alert_message(ip, component, status, life_remaining)
            else:
                print('')

        if data:
            show_alert(ip, data)
            printer_ips.append(ip)

        print('\n')

        return component_data

    except requests.exceptions.RequestException as e:
        print(f"Error connecting to {url}: {e}\nPrinter is offline or not found in the area.")

    except Exception as e:
        print(f"Error occurred in get_toner_data for {ip}: {e}")

        # Vuelve a ejecutar la función para intentar nuevamente
        return get_toner_data(url, ip, printer_ips)


def close_alert_windows():
    global alert_windows

    for window in alert_windows:
        window.destroy()
def show_email_alert(ip, alert_message):
    # Mostrar la alerta en la consola
    print(alert_message)

    # Verificar si la alerta ya existe en el archivo TXT
    alert_key = f"{ip}-email-alert"  # Clave única para la alerta de email
    if alert_key in existing_alerts:
        print("Alerta de email ya existe en el archivo TXT, no se enviará por correo electrónico.")
        return

    # Enviar la alerta por correo electrónico
    subject = f"ALERT: Printer at IP {ip}"
    send_email(subject, alert_message)

    # Guardar la alerta de email en la lista de alertas previamente guardadas
    existing_alerts.append(alert_key)

# Función para revisar las IP de la junta y enviar alertas
def check_ips_thread():
    while True:
        current_time = datetime.datetime.now().time()
        current_weekday = datetime.datetime.now().weekday()

        if current_weekday < 5 and current_time >= datetime.time(8) and current_time <= datetime.time(17):
            printer_ips = []
            printer_ips2 = []
            alert_windows.clear()
            for url, ip in urls:
                try:
                    response = requests.get(url, verify=False, timeout=10)
                    get_toner_data(url, ip, printer_ips)
                except requests.exceptions.RequestException:
                    alert_message = f"La impresora en la dirección IP {ip} está apagada o dejó de funcionar."
                    show_email_alert(ip, alert_message)

            printer_ips2 = urls2.copy()
            show_printer_ips(printer_ips2, window)

        time.sleep(5)  # Esperar 2 horas en segundos


def close_program():
    global ips_window, alert_windows

    close_alert_windows()  # Cerrar todas las ventanas de alerta abiertas
    close_ips_window()  # Cerrar la ventana de las IPs
    if os.path.exists("alertas.txt"):
        os.remove("alertas.txt")
    window.destroy()  # Cierra la ventana principal


# Crear la ventana principal
window = tk.Tk()
window.title("Printer Alert")
window.geometry("200x200")

# Iniciar la función check_ips en un hilo separado
check_ips_thread = threading.Thread(target=check_ips_thread)
check_ips_thread.daemon = True
check_ips_thread.start()


window.mainloop()